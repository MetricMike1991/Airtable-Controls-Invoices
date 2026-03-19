"""
report_generator.py – Generate Excel reports from Airtable view data.

Produces .xlsx files with:
  • All record fields as columns
  • Clickable "View in Airtable" links (using shared view URL)
  • Direct attachment download URLs for linked invoices
  • Summary row with totals
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
LINK_FONT = Font(name="Calibri", size=11, color="2563EB", underline="single")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
)


def generate_bank_txn_report(
    records: list[dict],
    linked_invoices: dict[str, dict],
    shared_view_url: str,
    output_path: str | Path,
    report_title: str = "Bank Transactions Report",
) -> Path:
    """
    Generate an Excel report for bank transaction records.

    Args:
        records: Raw Airtable records [{id, fields}, ...]
        linked_invoices: {invoice_record_id: {id, fields}} for linked invoices
        shared_view_url: Base shared view URL (e.g. https://airtable.com/shrXXX)
        output_path: Where to save the .xlsx file
        report_title: Title for the report header

    Returns: Path to the generated file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Transactions"

    # --- Title row ---
    ws.merge_cells("A1:H1")
    ws["A1"] = report_title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Headers ---
    headers = [
        "Date", "Description", "Amount", "Type", "Source",
        "Matched Invoice", "Invoice Link", "Attachment",
    ]
    col_widths = [14, 35, 14, 10, 10, 30, 20, 40]

    for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[3].height = 24

    # --- Data rows ---
    row_num = 4
    total_debit = 0.0
    total_credit = 0.0

    for rec in records:
        f = rec.get("fields", {})
        rec_id = rec.get("id", "")

        date_val = str(f.get("Date", ""))
        desc_val = str(f.get("Description", ""))
        amount_val = float(f.get("Amount", 0) or 0)
        type_val = str(f.get("Type", "Debit"))
        source_val = str(f.get("Source", ""))

        if type_val == "Debit":
            total_debit += amount_val
        else:
            total_credit += amount_val

        # Linked invoice info
        matched_ids = f.get("Matched Invoice", [])
        if isinstance(matched_ids, list):
            inv_ids = [x if isinstance(x, str) else x.get("id", "") for x in matched_ids]
        else:
            inv_ids = []

        inv_names = []
        inv_links = []
        attachment_urls = []

        for inv_id in inv_ids:
            inv_rec = linked_invoices.get(inv_id, {})
            inv_fields = inv_rec.get("fields", {})
            inv_name = str(inv_fields.get("Business Name", ""))
            inv_number = str(inv_fields.get("Invoice Number", ""))
            label = f"{inv_name} #{inv_number}" if inv_number else inv_name
            inv_names.append(label)

            # Build Airtable record link
            if shared_view_url:
                inv_links.append(f"{shared_view_url}/{inv_id}")

            # Attachment URLs
            attachments = inv_fields.get("Invoice Attachement", [])
            if isinstance(attachments, list):
                for att in attachments:
                    url = att.get("url", "")
                    if url:
                        attachment_urls.append(url)

        # Write cells
        ws.cell(row=row_num, column=1, value=date_val).font = DATA_FONT
        ws.cell(row=row_num, column=2, value=desc_val).font = DATA_FONT
        amt_cell = ws.cell(row=row_num, column=3, value=amount_val)
        amt_cell.font = DATA_FONT
        amt_cell.number_format = '€#,##0.00'
        ws.cell(row=row_num, column=4, value=type_val).font = DATA_FONT
        ws.cell(row=row_num, column=5, value=source_val).font = DATA_FONT

        # Matched invoice name(s)
        ws.cell(row=row_num, column=6, value=", ".join(inv_names) if inv_names else "—").font = DATA_FONT

        # Invoice link (clickable)
        if inv_links:
            link_cell = ws.cell(row=row_num, column=7, value="View Invoice")
            link_cell.hyperlink = inv_links[0]
            link_cell.font = LINK_FONT
        else:
            ws.cell(row=row_num, column=7, value="—").font = DATA_FONT

        # Attachment download link
        if attachment_urls:
            att_cell = ws.cell(row=row_num, column=8, value="Download")
            att_cell.hyperlink = attachment_urls[0]
            att_cell.font = LINK_FONT
        else:
            ws.cell(row=row_num, column=8, value="—").font = DATA_FONT

        # Row border
        for c in range(1, 9):
            ws.cell(row=row_num, column=c).border = BORDER

        row_num += 1

    # --- Summary row ---
    row_num += 1
    ws.cell(row=row_num, column=1, value="TOTALS").font = TOTAL_FONT
    ws.cell(row=row_num, column=1).fill = TOTAL_FILL

    ws.cell(row=row_num, column=2, value=f"{len(records)} records").font = TOTAL_FONT
    ws.cell(row=row_num, column=2).fill = TOTAL_FILL

    debit_cell = ws.cell(row=row_num, column=3, value=total_debit)
    debit_cell.font = TOTAL_FONT
    debit_cell.fill = TOTAL_FILL
    debit_cell.number_format = '€#,##0.00'

    ws.cell(row=row_num, column=4, value="Debit Total").font = TOTAL_FONT
    ws.cell(row=row_num, column=4).fill = TOTAL_FILL

    for c in range(5, 9):
        ws.cell(row=row_num, column=c).fill = TOTAL_FILL

    row_num += 1
    credit_cell = ws.cell(row=row_num, column=3, value=total_credit)
    credit_cell.font = TOTAL_FONT
    credit_cell.fill = TOTAL_FILL
    credit_cell.number_format = '€#,##0.00'

    ws.cell(row=row_num, column=4, value="Credit Total").font = TOTAL_FONT
    ws.cell(row=row_num, column=4).fill = TOTAL_FILL

    for c in [1, 2, 5, 6, 7, 8]:
        ws.cell(row=row_num, column=c).fill = TOTAL_FILL

    # --- Save ---
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def generate_invoice_report(
    records: list[dict],
    linked_bank_txns: dict[str, dict],
    shared_view_url: str,
    output_path: str | Path,
    report_title: str = "Invoices Report",
) -> Path:
    """
    Generate an Excel report for invoice records.

    Args:
        records: Raw Airtable records [{id, fields}, ...]
        linked_bank_txns: {bank_txn_id: {id, fields}} for linked bank transactions
        shared_view_url: Base shared view URL
        output_path: Where to save the .xlsx
        report_title: Title for the report header

    Returns: Path to the generated file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # --- Title row ---
    ws.merge_cells("A1:H1")
    ws["A1"] = report_title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Headers ---
    headers = [
        "Business Name", "Invoice #", "Date", "Total (inc VAT)",
        "Matched Bank Txn", "Bank Txn Link", "Attachment", "Notes",
    ]
    col_widths = [30, 16, 14, 16, 30, 20, 40, 30]

    for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[3].height = 24

    # --- Data rows ---
    row_num = 4
    total_amount = 0.0

    for rec in records:
        f = rec.get("fields", {})
        rec_id = rec.get("id", "")

        biz_name = str(f.get("Business Name", ""))
        inv_number = str(f.get("Invoice Number", ""))
        inv_date = str(f.get("Date Of Invoice", ""))
        notes = str(f.get("Additional Notes", ""))

        total_raw = f.get("Total Invoice Including VAT", "")
        try:
            total_val = float(
                str(total_raw).replace(",", "").replace("€", "").replace("£", "").replace("$", "").strip()
            )
        except (ValueError, TypeError):
            total_val = 0.0
        total_amount += total_val

        # Linked bank transaction info
        linked_ids = f.get("Bank Transactions", [])
        if isinstance(linked_ids, list):
            txn_ids = [x if isinstance(x, str) else x.get("id", "") for x in linked_ids]
        else:
            txn_ids = []

        txn_descs = []
        for tid in txn_ids:
            txn_rec = linked_bank_txns.get(tid, {})
            txn_fields = txn_rec.get("fields", {})
            desc = str(txn_fields.get("Description", ""))
            amt = txn_fields.get("Amount", "")
            label = f"{desc} (€{amt})" if amt else desc
            txn_descs.append(label)

        # Attachment URLs
        attachment_urls = []
        attachments = f.get("Invoice Attachement", [])
        if isinstance(attachments, list):
            for att in attachments:
                url = att.get("url", "")
                if url:
                    attachment_urls.append(url)

        # Write cells
        ws.cell(row=row_num, column=1, value=biz_name).font = DATA_FONT
        ws.cell(row=row_num, column=2, value=inv_number).font = DATA_FONT
        ws.cell(row=row_num, column=3, value=inv_date).font = DATA_FONT

        total_cell = ws.cell(row=row_num, column=4, value=total_val)
        total_cell.font = DATA_FONT
        total_cell.number_format = '€#,##0.00'

        ws.cell(row=row_num, column=5, value=", ".join(txn_descs) if txn_descs else "—").font = DATA_FONT

        # Bank txn link
        if txn_ids and shared_view_url:
            link_cell = ws.cell(row=row_num, column=6, value="View Transaction")
            link_cell.hyperlink = f"{shared_view_url}/{txn_ids[0]}"
            link_cell.font = LINK_FONT
        else:
            ws.cell(row=row_num, column=6, value="—").font = DATA_FONT

        # Attachment download
        if attachment_urls:
            att_cell = ws.cell(row=row_num, column=7, value="Download")
            att_cell.hyperlink = attachment_urls[0]
            att_cell.font = LINK_FONT
        else:
            ws.cell(row=row_num, column=7, value="—").font = DATA_FONT

        ws.cell(row=row_num, column=8, value=notes).font = DATA_FONT

        # Row border
        for c in range(1, 9):
            ws.cell(row=row_num, column=c).border = BORDER

        row_num += 1

    # --- Summary row ---
    row_num += 1
    ws.cell(row=row_num, column=1, value="TOTALS").font = TOTAL_FONT
    ws.cell(row=row_num, column=1).fill = TOTAL_FILL

    ws.cell(row=row_num, column=2, value=f"{len(records)} invoices").font = TOTAL_FONT
    ws.cell(row=row_num, column=2).fill = TOTAL_FILL

    for c in [3]:
        ws.cell(row=row_num, column=c).fill = TOTAL_FILL

    total_cell = ws.cell(row=row_num, column=4, value=total_amount)
    total_cell.font = TOTAL_FONT
    total_cell.fill = TOTAL_FILL
    total_cell.number_format = '€#,##0.00'

    for c in range(5, 9):
        ws.cell(row=row_num, column=c).fill = TOTAL_FILL

    # --- Save ---
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
